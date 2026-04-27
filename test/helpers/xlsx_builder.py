"""In-memory XLSX builders for import integration tests.

Generates files in the exact format expected by the Go API importer.
"""
import io
from datetime import datetime, timedelta, timezone

import openpyxl


def build_stammdaten_xlsx(rows: list[dict]) -> bytes:
    """Build a Stammdaten XLSX in the format expected by ParseStammdaten.

    Sheet layout (0-based row indices):
      0-5  : skipped header / marker rows
      6    : column headers (at the exact column positions the parser uses)
      7-8  : skipped marker rows
      9+   : data rows (only ACTIVATED rows are imported)

    Row dict keys (all optional except zaehlpunkt, name1, mitglieds_nr):
      netzbetreiber, gemeinschaft_id, zaehlpunkt, energierichtung,
      verteilungsmodell, zugeteilte_menge_pct, name1, name2,
      business_role, iban, email, mitglieds_nr, status, registriert_seit
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EEG Stammdaten"

    # Rows 1-6 (0-based 0-5): dummy skip rows
    for i in range(1, 7):
        ws.cell(row=i, column=1, value=f"Skip row {i}")

    # Row 7 (0-based 6): column headers at the exact parser column positions
    # Parser uses 0-based indices; openpyxl uses 1-based columns.
    _hdr = 7
    ws.cell(row=_hdr, column=1,  value="Netzbetreiber")          # col 0
    ws.cell(row=_hdr, column=2,  value="Gemeinschaft ID")         # col 1
    ws.cell(row=_hdr, column=12, value="Zählpunkt")               # col 11
    ws.cell(row=_hdr, column=13, value="Energierichtung")         # col 12
    ws.cell(row=_hdr, column=18, value="Verteilungsmodell")       # col 17
    ws.cell(row=_hdr, column=19, value="Zugeteilte Menge %")      # col 18
    ws.cell(row=_hdr, column=21, value="Name 1")                  # col 20
    ws.cell(row=_hdr, column=22, value="Name 2")                  # col 21
    ws.cell(row=_hdr, column=24, value="Geschäftsrolle")          # col 23
    ws.cell(row=_hdr, column=25, value="IBAN")                    # col 24
    ws.cell(row=_hdr, column=27, value="E-Mail")                  # col 26
    ws.cell(row=_hdr, column=29, value="Mitglieds-Nr")            # col 28
    ws.cell(row=_hdr, column=30, value="Zählpunktstatus")         # col 29
    ws.cell(row=_hdr, column=32, value="Registriert seit")        # col 31

    # Rows 8-9 (0-based 7-8): marker rows (skipped by parser)
    ws.cell(row=8, column=1, value="Marker 1")
    ws.cell(row=9, column=1, value="Marker 2")

    # Data rows start at row 10 (0-based index 9)
    for idx, row in enumerate(rows):
        r = 10 + idx
        ws.cell(row=r, column=1,  value=row.get("netzbetreiber", ""))
        ws.cell(row=r, column=2,  value=row.get("gemeinschaft_id", ""))
        ws.cell(row=r, column=12, value=row.get("zaehlpunkt", ""))
        ws.cell(row=r, column=13, value=row.get("energierichtung", "CONSUMPTION"))
        ws.cell(row=r, column=18, value=row.get("verteilungsmodell", "DYNAMIC"))
        ws.cell(row=r, column=19, value=float(row.get("zugeteilte_menge_pct", 100)))
        ws.cell(row=r, column=21, value=row.get("name1", ""))
        ws.cell(row=r, column=22, value=row.get("name2", ""))
        ws.cell(row=r, column=24, value=row.get("business_role", "privat"))
        ws.cell(row=r, column=25, value=row.get("iban", ""))
        ws.cell(row=r, column=27, value=row.get("email", ""))
        ws.cell(row=r, column=29, value=row.get("mitglieds_nr", ""))
        ws.cell(row=r, column=30, value=row.get("status", "ACTIVATED"))
        ws.cell(row=r, column=32, value=row.get("registriert_seit", "2025-01-01"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_energiedaten_xlsx(
    meter_ids: list[str],
    period_start: datetime,
    hours: int = 24,
    wh_total: float = 500.0,
    wh_community: float = 300.0,
) -> bytes:
    """Build an Energiedaten XLSX in the format expected by ParseEnergieDaten.

    Sheet layout:
      Row 1: "MeteringPoint"  | meter_id  | meter_id  | ...  (2 cols per meter)
      Row 2: "Metercode"      | col_hdr1  | col_hdr2  | ...
      Row 3+: "YYYY-MM-DD HH:MM:SS" | val | val | ...

    Two columns per meter: Gesamtverbrauch (total) + Anteil gemeinschaftliche (community).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Energiedaten"

    # Column layout: col 1 = timestamp, then 2 cols per meter
    meter_cols: dict[str, tuple[int, int]] = {}
    col = 2
    for mid in meter_ids:
        meter_cols[mid] = (col, col + 1)
        col += 2

    # Row 1: MeteringPoint row
    ws.cell(row=1, column=1, value="MeteringPoint")
    for mid, (tc, cc) in meter_cols.items():
        ws.cell(row=1, column=tc, value=mid)
        ws.cell(row=1, column=cc, value=mid)

    # Row 2: Metercode row (column type headers)
    ws.cell(row=2, column=1, value="Metercode")
    for _, (tc, cc) in meter_cols.items():
        ws.cell(row=2, column=tc, value="Gesamtverbrauch")
        ws.cell(row=2, column=cc, value="Anteil gemeinschaftliche Erzeugung")

    # Data rows
    for h in range(hours):
        ts = period_start + timedelta(hours=h)
        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
        r = 3 + h
        ws.cell(row=r, column=1, value=ts_str)
        for _, (tc, cc) in meter_cols.items():
            ws.cell(row=r, column=tc, value=wh_total)
            ws.cell(row=r, column=cc, value=wh_community)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
