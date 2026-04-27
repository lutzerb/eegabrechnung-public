"""Accounting export tests: XLSX Buchungsjournal + DATEV CSV.

Tests GET /api/v1/eegs/{eegID}/accounting/export?from=…&to=…&format=xlsx|datev

Requires test_readings fixture (provides readings for Jan–May 2025).
Uses April 2025 as the billing period (readings exist) but the accounting export
uses the CURRENT month range — invoices are created today (created_at = now()),
so the export must query by the current date, not the billing period dates.
"""
import zipfile
from datetime import date

import pytest

from helpers.api import APIError

# Billing run uses a period with readings (conftest inserts Jan–May 2025).
BILLING_FROM = "2025-04-01"
BILLING_TO   = "2025-04-30"

# Accounting export query: invoices are created NOW, so query the current month.
_today = date.today()
EXPORT_FROM = _today.replace(day=1).isoformat()
EXPORT_TO   = _today.isoformat()


class TestAccountingExport:
    """Full cycle: finalize a billing run, then export in both formats."""

    @pytest.fixture(scope="class")
    def finalized_run(self, api, test_eeg, test_readings):
        """Create + finalize a billing run for Apr 2025."""
        eeg_id = test_eeg["id"]
        run = api.create_billing_run(eeg_id, {
            "period_start": BILLING_FROM,
            "period_end": BILLING_TO,
        })
        api.finalize_billing_run(eeg_id, run["id"])
        yield {"eeg_id": eeg_id, "run_id": run["id"]}
        # Cleanup: cancel so the period can be reused
        try:
            api.cancel_billing_run(eeg_id, run["id"])
        except Exception:
            pass

    def test_xlsx_export_is_valid_xlsx(self, api, finalized_run):
        data = api.accounting_export(
            finalized_run["eeg_id"], EXPORT_FROM, EXPORT_TO, fmt="xlsx"
        )
        assert len(data) > 0, "Response is empty"
        # XLSX files are ZIP archives starting with PK
        assert data[:2] == b"PK", "Response is not a valid XLSX (missing PK header)"
        # Verify it's a proper ZIP/XLSX by opening it
        import io
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = zf.namelist()
        assert any("xl/worksheets" in n for n in names), (
            f"No worksheets found in XLSX archive: {names}"
        )

    def test_xlsx_export_contains_buchungsjournal_sheet(self, api, finalized_run):
        import io
        import openpyxl
        data = api.accounting_export(
            finalized_run["eeg_id"], BILLING_FROM, BILLING_TO, fmt="xlsx"
        )
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Buchungsjournal" in wb.sheetnames, (
            f"Expected sheet 'Buchungsjournal', found: {wb.sheetnames}"
        )
        ws = wb["Buchungsjournal"]
        # Header row should have expected columns
        headers = [ws.cell(row=1, column=c).value for c in range(1, 15)]
        assert "Belegdatum" in headers
        assert "Bruttobetrag EUR" in headers

    def test_datev_export_starts_with_extf(self, api, finalized_run):
        data = api.accounting_export(
            finalized_run["eeg_id"], BILLING_FROM, BILLING_TO, fmt="datev"
        )
        text = data.decode("utf-8", errors="replace")
        assert text.startswith('"EXTF"'), (
            f"DATEV export must start with '\"EXTF\"', got: {text[:80]!r}"
        )

    def test_datev_export_has_header_and_column_row(self, api, finalized_run):
        data = api.accounting_export(
            finalized_run["eeg_id"], BILLING_FROM, BILLING_TO, fmt="datev"
        )
        lines = data.decode("utf-8", errors="replace").splitlines()
        assert len(lines) >= 2, "DATEV export must have at least 2 header rows"
        # Row 2 must contain DATEV column headers
        assert "Umsatz" in lines[1], (
            f"Row 2 should contain column headers, got: {lines[1][:120]!r}"
        )

    def test_datev_export_has_data_rows_when_invoices_exist(self, api, finalized_run):
        # Use current-month range: invoices created_at = now (not the billing period).
        data = api.accounting_export(
            finalized_run["eeg_id"], EXPORT_FROM, EXPORT_TO, fmt="datev"
        )
        lines = [l for l in data.decode("utf-8", errors="replace").splitlines() if l.strip()]
        # Lines 1+2 are header; line 3+ are data rows
        assert len(lines) >= 3, (
            "Expected at least one data row in DATEV export (invoices must exist). "
            f"Export range: {EXPORT_FROM} – {EXPORT_TO}"
        )

    def test_missing_from_date_returns_400(self, api, finalized_run):
        with pytest.raises(APIError) as exc:
            api.accounting_export(finalized_run["eeg_id"], "", BILLING_TO)
        assert exc.value.status_code == 400

    def test_missing_to_date_returns_400(self, api, finalized_run):
        with pytest.raises(APIError) as exc:
            api.accounting_export(finalized_run["eeg_id"], BILLING_FROM, "")
        assert exc.value.status_code == 400

    def test_export_empty_period_returns_valid_files(self, api, test_eeg, test_readings):
        """Period with no invoices should still return a valid (empty) file."""
        eeg_id = test_eeg["id"]
        # Use a year with no invoices
        xlsx = api.accounting_export(eeg_id, "2020-01-01", "2020-12-31", fmt="xlsx")
        assert xlsx[:2] == b"PK", "Empty XLSX should still be valid"

        datev = api.accounting_export(eeg_id, "2020-01-01", "2020-12-31", fmt="datev")
        text = datev.decode("utf-8", errors="replace")
        assert text.startswith('"EXTF"'), "Empty DATEV should still have header"
